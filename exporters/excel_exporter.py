"""
Excel exporter.

Exports processed HSE Training Records into a formatted Excel workbook.
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from models.batch_result import BatchResult


class ExcelExporter:
    """Exports BatchResult objects to Excel."""

    HEADER_FILL = PatternFill(
        fill_type="solid",
        fgColor="1F4E78",
    )

    HEADER_FONT = Font(
        color="FFFFFF",
        bold=True,
    )

    THIN_BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    COLUMNS = [

        (
            "Training ID",
            lambda result, index:
            f"TRN-{index:06d}",
        ),

        (
            "Filename",
            lambda result, index:
            result.pdf_path.name,
        ),

        (
            "Details (Subject)",
            lambda result, index:
            result.training_record.training_subject
            if result.training_record else "",
        ),

        (
            "Date",
            lambda result, index:
            result.training_record.date
            if result.training_record else "",
        ),

        (
            "Number of Attendees",
            lambda result, index:
            result.training_record.number_of_attendees
            if result.training_record else "",
        ),

        (
            "Type",
            lambda result, index:
            result.training_record.document_type.value
            if result.training_record else "",
        ),

        (
            "Topic (Hazard Category)",
            lambda result, index:
            ", ".join(result.training_record.hazard_categories)
            if result.training_record else "",
        ),

        (
            "Trainer",
            lambda result, index:
            result.training_record.trainer
            if result.training_record else "",
        ),

        (
            "Duration",
            lambda result, index:
            result.training_record.duration_text
            if result.training_record else "",
        ),

    ]

    def export(
        self,
        batch_result: BatchResult,
        output_file: Path,
    ) -> None:
        """Export the batch results."""

        workbook = Workbook()

        training_sheet = workbook.active
        training_sheet.title = "Training Records"

        summary_sheet = workbook.create_sheet(
            "Processing Summary"
        )

        self._create_training_sheet(
            training_sheet,
            batch_result,
        )

        self._create_summary_sheet(
            summary_sheet,
            batch_result,
        )

        workbook.save(Path(output_file))

    def _create_training_sheet(
        self,
        worksheet,
        batch_result: BatchResult,
    ) -> None:
        """Create the Training Records worksheet."""

        # Header row
        for column, (heading, _) in enumerate(self.COLUMNS, start=1):

            cell = worksheet.cell(
                row=1,
                column=column,
                value=heading,
            )

            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.border = self.THIN_BORDER

        # Data rows
        row = 2

        for index, result in enumerate(
            batch_result.successful_results,
            start=1,
        ):

            for column, (_, getter) in enumerate(
                self.COLUMNS,
                start=1,
            ):

                value = getter(result, index)

                cell = worksheet.cell(
                    row=row,
                    column=column,
                    value=value,
                )

                cell.border = self.THIN_BORDER

                # Date formatting
                if column == 4 and value:
                    cell.number_format = "DD/MM/YYYY"

            row += 1

        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        self._auto_size_columns(worksheet)

    def _create_summary_sheet(
        self,
        worksheet,
        batch_result: BatchResult,
    ) -> None:
        """Create the Processing Summary worksheet."""

        worksheet["A1"] = "HSE Document Processor"
        worksheet["A1"].font = Font(
            bold=True,
            size=16,
        )

        summary = [

            ("Generated", datetime.now().strftime("%d %b %Y %H:%M")),

            ("Documents Processed",
             batch_result.total_documents),

            ("Successful",
             batch_result.successful_documents),

            ("Failed",
             batch_result.failed_documents),

            ("Success Rate",
             f"{batch_result.success_rate:.2f}%"),

            ("Average Processing Time",
             f"{batch_result.average_processing_time:.2f} sec"),

            ("Total Processing Time",
             f"{batch_result.processing_time_seconds:.2f} sec"),

        ]

        row = 3

        for label, value in summary:

            label_cell = worksheet.cell(
                row=row,
                column=1,
                value=label,
            )

            value_cell = worksheet.cell(
                row=row,
                column=2,
                value=value,
            )

            label_cell.font = Font(bold=True)

            label_cell.border = self.THIN_BORDER
            value_cell.border = self.THIN_BORDER

            row += 1

        self._auto_size_columns(worksheet)

    def _auto_size_columns(
        self,
        worksheet,
    ) -> None:
        """Automatically size worksheet columns."""

        for column_cells in worksheet.columns:

            length = max(
                len(str(cell.value))
                if cell.value is not None else 0
                for cell in column_cells
            )

            worksheet.column_dimensions[
                get_column_letter(
                    column_cells[0].column
                )
            ].width = min(length + 4, 50)